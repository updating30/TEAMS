"""
Teams会議文字起こし（.vtt） → Excel自動記入スクリプト
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

CONFIG_FILE  = Path("config/cell_mapping.json")
INPUT_DIR    = Path("input")
OUTPUT_DIR   = Path("output")
TEMPLATE_DIR = Path("template")

TRIGGER     = "記入"
END_MARKER  = "以上"

# セルアドレス判定（例: C3, AA10）
_CELL_ADDR_RE = re.compile(r"^[A-Za-z]+\d+$")
# 区切り文字（読点・句点・カンマ・空白）
_SPLIT_RE = re.compile(r"[、，,。\s　]+")


# ---------------------------------------------------------------------------
# 設定読み込み
# ---------------------------------------------------------------------------

def load_config() -> dict:
    if not CONFIG_FILE.exists():
        print(f"エラー: {CONFIG_FILE} が見つかりません")
        sys.exit(1)
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# VTT パーサー（ブロック単位で返す。結合しない）
# ---------------------------------------------------------------------------

def parse_vtt(filepath: Path) -> list[dict]:
    """
    Teams .vtt を解析し {'speaker': str, 'text': str} のリストを返す。
    ブロックは結合せず、VTT のキュー単位で1要素とする。
    """
    with open(filepath, encoding="utf-8") as f:
        content = f.read()

    messages: list[dict] = []
    for block in re.split(r"\n\n+", content.strip()):
        lines = [l.rstrip() for l in block.splitlines() if l.strip()]
        if not lines or lines[0].strip() == "WEBVTT":
            continue

        ts_idx = next((i for i, l in enumerate(lines) if "-->" in l), None)
        if ts_idx is None:
            continue

        text_lines = lines[ts_idx + 1:]
        if not text_lines:
            continue

        full = " ".join(text_lines)
        speaker = ""

        # <v 話者名>テキスト 形式
        m = re.match(r"<v ([^>]+)>(.*)", full, re.DOTALL)
        if m:
            speaker = m.group(1).strip()
            text    = re.sub(r"<[^>]+>", "", m.group(2)).strip()
        else:
            # 話者名: テキスト 形式
            m2 = re.match(r"^([^:：\n]{1,30})[：:]\s*(.*)", full, re.DOTALL)
            if m2:
                speaker = m2.group(1).strip()
                text    = re.sub(r"<[^>]+>", "", m2.group(2)).strip()
            else:
                text = re.sub(r"<[^>]+>", "", full).strip()

        if text:
            messages.append({"speaker": speaker, "text": text})

    return messages


def parse_text_file(filepath: Path) -> list[dict]:
    """テスト用：テキストファイルを行単位で読む。"""
    messages = []
    for line in filepath.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            messages.append({"speaker": "", "text": line})
    return messages


# ---------------------------------------------------------------------------
# コマンド抽出
# ---------------------------------------------------------------------------

def _split_parts(text: str) -> list[str]:
    return [p for p in _SPLIT_RE.split(text) if p]


def parse_commands(messages: list[dict]) -> list[dict]:
    """
    発話リストから「記入」コマンドを抽出する。

    ルール：
    - 名前付きフィールド（担当者 等）は「記入」を含むブロック内で完結する
    - 連続数値データ（C5 等のセルアドレス）は「以上」が出るまで後続ブロックを引き継ぐ
    """
    commands: list[dict] = []
    # 連続データの引き継ぎ用
    pending_parts:   list[str] | None = None
    pending_speaker: str | None       = None

    for msg in messages:
        text    = msg["text"]
        speaker = msg["speaker"]

        # ---- 連続データの引き継ぎ中 ----
        if pending_parts is not None:
            if END_MARKER in text:
                # 「以上」が来たら確定
                idx_end = text.find(END_MARKER)
                extra   = _split_parts(text[:idx_end])
                pending_parts.extend(v for v in extra if v != END_MARKER)
                commands.append({"speaker": pending_speaker, "parts": list(pending_parts)})
                pending_parts   = None
                pending_speaker = None
                # 「以上」より後ろに新しい「記入」があれば続けて処理
                text = text[idx_end + len(END_MARKER):]
                if TRIGGER not in text:
                    continue
                # fall through → 通常処理
            elif TRIGGER not in text:
                # 「記入」も「以上」もない＝数値の続き
                extra = _split_parts(text)
                pending_parts.extend(extra)
                continue
            else:
                # 新しい「記入」が来た＝pending を確定して通常処理へ
                commands.append({"speaker": pending_speaker, "parts": list(pending_parts)})
                pending_parts   = None
                pending_speaker = None
                # fall through → 通常処理

        # ---- 通常処理：「記入」トリガーを探す ----
        if TRIGGER not in text:
            continue

        start = 0
        while True:
            idx = text.find(TRIGGER, start)
            if idx == -1:
                break

            after = text[idx + len(TRIGGER):].strip()

            # 次の「記入」までを切り出す（同一発話内に複数コマンドがある場合）
            next_trigger = after.find(TRIGGER)
            segment = after[:next_trigger] if next_trigger != -1 else after

            # 「以上」で打ち切る
            end_idx  = segment.find(END_MARKER)
            has_end  = end_idx != -1
            if has_end:
                segment = segment[:end_idx]

            parts = _split_parts(segment)
            # parts[0]=シート識別子  parts[1]=フィールド名orセルアドレス  parts[2:]=値
            if len(parts) < 3:
                start = idx + len(TRIGGER) + 1
                continue

            is_cell = bool(_CELL_ADDR_RE.match(parts[1]))

            if is_cell and not has_end:
                # 「以上」なしの連続データ → 後続ブロックへ引き継ぎ
                pending_parts   = parts
                pending_speaker = speaker
            else:
                commands.append({"speaker": speaker, "parts": parts})

            start = idx + len(TRIGGER) + 1

    # ファイル末尾まで読んで pending が残っていれば確定
    if pending_parts is not None:
        commands.append({"speaker": pending_speaker, "parts": list(pending_parts)})

    return commands


# ---------------------------------------------------------------------------
# Excel への書き込み
# ---------------------------------------------------------------------------

def _cell_to_col_row(cell_addr: str):
    """'C3' → (col=3, row=3)"""
    m = re.match(r"([A-Za-z]+)(\d+)$", cell_addr)
    if not m:
        return None, None
    return column_index_from_string(m.group(1).upper()), int(m.group(2))


def _to_number(val: str):
    """数値変換できれば変換、できなければ文字列のまま返す。"""
    cleaned = val.replace(",", "").replace("，", "").replace(" ", "")
    try:
        f = float(cleaned)
        return int(f) if f == int(f) else f
    except ValueError:
        return val


def _resolve_sheet(
    wb: openpyxl.Workbook,
    sheet_cfg: dict,
) -> openpyxl.worksheet.worksheet.Worksheet:
    name = sheet_cfg.get("excel_sheet_name", "")
    if name and name in wb.sheetnames:
        return wb[name]
    return wb.active


def fill_excel(
    template_path: Path,
    commands: list[dict],
    config: dict,
    output_path: Path,
) -> None:
    wb = openpyxl.load_workbook(template_path)
    sheets_config: dict = config.get("sheets", {})

    for cmd in commands:
        parts   = cmd["parts"]
        speaker = cmd["speaker"]

        if len(parts) < 3:
            print(f"  スキップ（要素不足）: {parts}")
            continue

        raw_key       = parts[0]
        field_or_cell = parts[1]
        values        = parts[2:]

        # 「シートAAA」→「AAA」正規化
        normalized_key = re.sub(r"^シート", "", raw_key)
        sheet_cfg = sheets_config.get(raw_key) or sheets_config.get(normalized_key)

        if sheet_cfg is None:
            print(f"  警告: シート '{raw_key}' が設定ファイルに見つかりません")
            continue

        ws = _resolve_sheet(wb, sheet_cfg)

        if not _CELL_ADDR_RE.match(field_or_cell):
            # ---- 名前付きフィールド ----
            fields: dict = sheet_cfg.get("fields", {})
            if field_or_cell not in fields:
                print(f"  警告: フィールド '{field_or_cell}' が設定にありません")
                continue
            cell_addr = fields[field_or_cell]
            # 値は parts[2] のみ（1フィールドに1値）
            value = values[0]
            ws[cell_addr] = value
            print(f"  書込 [{speaker}]: {raw_key}/{field_or_cell} → {cell_addr} = {value!r}")

        else:
            # ---- 連続データ ----
            col, row = _cell_to_col_row(field_or_cell)
            if col is None:
                print(f"  警告: セルアドレス '{field_or_cell}' を解析できません")
                continue

            # 書き込み方向（設定ファイルから取得、デフォルト右）
            direction = "right"
            for rng in sheet_cfg.get("ranges", {}).values():
                if rng.get("start_cell", "").upper() == field_or_cell.upper():
                    direction = rng.get("direction", "right")
                    break

            written = 0
            for i, val in enumerate(values):
                converted = _to_number(val)
                if direction == "right":
                    ws.cell(row=row, column=col + i, value=converted)
                else:
                    ws.cell(row=row + i, column=col, value=converted)
                written += 1

            print(f"  書込 [{speaker}]: {raw_key}/{field_or_cell} 方向={direction} → {written}個")

    wb.save(output_path)
    print(f"  -> 保存完了: {output_path}")


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main() -> None:
    config = load_config()

    templates = sorted(TEMPLATE_DIR.glob("*.xlsx"))
    if not templates:
        print("エラー: template/ フォルダに .xlsx ファイルが見つかりません")
        sys.exit(1)
    template_path = templates[0]
    print(f"テンプレート: {template_path}\n")

    input_files = sorted(INPUT_DIR.glob("*.vtt")) + sorted(INPUT_DIR.glob("*.txt"))
    if not input_files:
        print("エラー: input/ フォルダに .vtt または .txt ファイルがありません")
        print("       Teams文字起こしファイルを input/ フォルダに置いてください")
        sys.exit(1)

    for input_file in input_files:
        print(f"処理中: {input_file.name}")
        if input_file.suffix.lower() == ".vtt":
            messages = parse_vtt(input_file)
        else:
            messages = parse_text_file(input_file)

        commands = parse_commands(messages)
        if not commands:
            print("  警告: '記入' コマンドが見つかりませんでした")
            continue

        print(f"  {len(commands)} 件のコマンドを検出")
        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"{input_file.stem}_{timestamp}.xlsx"
        fill_excel(template_path, commands, config, output_path)

    print("\n処理が完了しました。")


if __name__ == "__main__":
    main()
